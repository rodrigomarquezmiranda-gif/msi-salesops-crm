"""
MSI SalesOps — Stock & Price Check + Email Notification
Corre vía GitHub Actions todos los días hábiles a las 18:30 ART.
"""
import json, re, sys, os, io, smtplib
from datetime import datetime, timezone
from urllib import request as urlrequest
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# ── Config ────────────────────────────────────────────────────────────────────
FB_BASE    = "https://msi-crm-default-rtdb.firebaseio.com"
FB_SECRET  = os.environ.get("FIREBASE_SECRET", "")
FB_PARAMS  = f"?auth={FB_SECRET}" if FB_SECRET else ""

SMTP_HOST      = "smtp.hostinger.com"
SMTP_PORT      = 465
SMTP_USER      = "sales@msicrm.com"
SMTP_PASS      = os.environ.get("SMTP_PASSWORD", "")
FORCE_EMAIL    = os.environ.get("FORCE_EMAIL", "false").lower() == "true"
TEST_RECIPIENT = os.environ.get("TEST_RECIPIENT", "").strip()
_raw_recipients = TEST_RECIPIENT if TEST_RECIPIENT else os.environ.get("RECIPIENT_EMAILS", "")
RECIPIENTS = [e.strip() for e in _raw_recipients.split(",") if e.strip()]

# ── Firebase helpers ──────────────────────────────────────────────────────────
def fb_get(path):
    url = f"{FB_BASE}/{path}.json{FB_PARAMS}"
    try:
        with urlrequest.urlopen(url, timeout=30) as r:
            return json.loads(r.read())
    except Exception as e:
        print(f"[WARN] GET {path}: {e}"); return None

def fb_put(path, data):
    url = f"{FB_BASE}/{path}.json{FB_PARAMS}"
    req = urlrequest.Request(url, data=json.dumps(data).encode(),
                             method="PUT", headers={"Content-Type":"application/json"})
    try:
        with urlrequest.urlopen(req, timeout=30) as r:
            print(f"[OK]   PUT {path} → {r.status}"); return True
    except Exception as e:
        print(f"[ERR]  PUT {path}: {e}"); return False

def fb_post(path, data):
    url = f"{FB_BASE}/{path}.json{FB_PARAMS}"
    req = urlrequest.Request(url, data=json.dumps(data).encode(),
                             method="POST", headers={"Content-Type":"application/json"})
    try:
        with urlrequest.urlopen(req, timeout=30) as r:
            print(f"[OK]   POST {path} → {r.status}"); return True
    except Exception as e:
        print(f"[ERR]  POST {path}: {e}"); return False

def safe_int(v):
    try: m = re.match(r'(\d+)', str(v).strip()); return int(m.group(1)) if m else 0
    except: return 0

def safe_float(v):
    try: return float(str(v).replace('$','').replace(',','').strip())
    except: return 0.0

# ── Generate Excel — IDENTICAL to CRM exportPriceListExcel() ─────────────────
def generate_excel(raw_products, changes, date_str):
    """
    Replica exacta de exportPriceListExcel() en index.html.
    Mismos colores, fuentes, columnas, sub-secciones y formato de precio.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[WARN] openpyxl no disponible"); return None

    # ── CRM palette (argb sin el prefijo FF para openpyxl) ──
    NAVY    = "1F2A44"   # title bg
    NAVY2   = "30456B"   # date bg
    RED     = "C8102E"   # category header
    WHITE   = "FFFFFF"
    SUBSEC  = "DCE6F1"   # sub-section banner (light blue)
    ALT1    = "FFFFFF"   # alternating row 1
    ALT2    = "F7F9FC"   # alternating row 2
    GREEN_T = "1A5C28"   # price text color

    # Estado fill/text/bold exactly as in CRM
    ESTADO_FILL  = {"new!":"92D050","focus!":"4472C4","eol":"FF0000",
                    "to clean":"808080","last batch / eol soon":"FFC000",
                    "shortage":"FF66CC","wall plug!":"808080"}
    ESTADO_TEXT  = {"new!":WHITE,"focus!":WHITE,"eol":WHITE,"to clean":WHITE,
                    "last batch / eol soon":"000000","shortage":WHITE,"wall plug!":"000000"}
    ESTADO_BOLD  = {"new!":True,"focus!":True,"eol":True,"to clean":True,
                    "last batch / eol soon":True,"shortage":True,"wall plug!":True}

    BORDER = Border(
        top=Side(style="thin",color="FFCCCCCC"),
        left=Side(style="thin",color="FFCCCCCC"),
        bottom=Side(style="thin",color="FFCCCCCC"),
        right=Side(style="thin",color="FFCCCCCC"),
    )

    def sf(argb):
        return PatternFill("solid", fgColor=argb)

    def cfont(size=10, bold=False, italic=False, color=None, name="Calibri"):
        return Font(size=size, bold=bold, italic=italic,
                    color=color or "FF000000", name=name)

    NC = 13
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lista de Precios"

    # Column widths (px ≈ Excel units, from CRM source)
    for col, w in enumerate([16,20,40,28,12,11,9,13,12,20,14,22,13], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    def add_merged(val, fill_argb, font_obj, height=None):
        ws.append([""])
        rn = ws.max_row
        ws.merge_cells(start_row=rn, start_column=1, end_row=rn, end_column=NC)
        mc = ws.cell(rn, 1)
        if val is not None: mc.value = val
        if fill_argb: mc.fill = sf(fill_argb)
        if font_obj:  mc.font = font_obj
        mc.alignment = Alignment(horizontal="left", vertical="center")
        if height: ws.row_dimensions[rn].height = height

    # ── Rows 1-8: header block (identical to CRM) ──
    add_merged("MSI LATAM — Lista de Precios y Stock",
               NAVY, cfont(size=18, bold=True, color="FF"+WHITE), 31.5)
    add_merged(f"Fecha de referencia: {date_str}",
               NAVY2, cfont(size=11, color="FF"+WHITE), 15)
    add_merged(None, None, None, None)
    add_merged("Notas", None, cfont(size=10, bold=True), 15)
    nf = cfont(size=9, color="FF555555")
    add_merged("Miami Stock: físicamente en Miami.   Miami Bonded Stock: físicamente en Miami Bonded / FTZ.   ETA: llegada estimada.", None, nf, 15)
    add_merged("Precio: se toma el valor de la columna Miami Price; si existe un precio de Market Share menor, se usa el más bajo de los dos.", None, nf, 15)
    add_merged(f"Exportado desde SalesOps CRM · {date_str}", None, nf, 15)
    add_merged(None, None, None, None)

    # ── Row 9: column headers ──
    ws.append([""])
    hr = ws.max_row  # = 9
    for i, h in enumerate(["UPC","N° de Parte","Producto","Detalle","Estado","Links",
                            "Qty/Ctn","Qty/Plt","Miami Stock","Miami ETA",
                            "Miami Bonded Stock","Miami Bonded ETA","Precio"], 1):
        c = ws.cell(hr, i)
        c.value = h
        c.fill  = sf(NAVY)
        c.font  = cfont(size=10, bold=True, color="FF"+WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
    ws.row_dimensions[hr].height = 30

    # Freeze below header
    ws.freeze_panes = "A10"

    # ── Rows 10+: products ──
    # raw_products can be a list (Firebase array) or dict
    if isinstance(raw_products, list):
        items_list = [p for p in raw_products if isinstance(p, dict)]
    else:
        items_list = list(raw_products.values())

    current_cat = ""
    alt_idx = 0

    for item in items_list:
        if not isinstance(item, dict): continue

        # Sub-section banner (light blue)
        if item.get("isSubSection"):
            ws.append([""])
            srn = ws.max_row
            ws.merge_cells(start_row=srn, start_column=1, end_row=srn, end_column=NC)
            mc = ws.cell(srn, 1)
            mc.value = item.get("label","")
            mc.fill  = sf(SUBSEC)
            mc.font  = cfont(size=10, italic=True, color="FF"+NAVY)
            mc.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[srn].height = 15
            alt_idx = 0
            continue

        # Category header (red)
        cat = item.get("category","")
        if cat != current_cat:
            current_cat = cat
            alt_idx = 0
            ws.append([""])
            crn = ws.max_row
            ws.merge_cells(start_row=crn, start_column=1, end_row=crn, end_column=NC)
            mc = ws.cell(crn, 1)
            mc.value = cat
            mc.fill  = sf(RED)
            mc.font  = cfont(size=11, bold=True, color="FF"+WHITE)
            mc.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[crn].height = 21.75

        bg = ALT1 if alt_idx % 2 == 0 else ALT2
        alt_idx += 1

        ws.append([""])
        prn = ws.max_row
        ws.row_dimensions[prn].height = 15

        def setc(col, val, font_obj=None, num_fmt=None, halign="left"):
            c = ws.cell(prn, col)
            c.value = val
            c.fill  = sf(bg)
            c.border = BORDER
            c.font  = font_obj or cfont(size=10)
            c.alignment = Alignment(horizontal=halign, vertical="center")
            if num_fmt: c.number_format = num_fmt

        # Col 1: UPC (as number)
        upc_raw = item.get("upc","")
        upc_val = int(upc_raw) if str(upc_raw).isdigit() else (upc_raw or "")
        setc(1, upc_val, cfont(size=10),
             "0" if isinstance(upc_val, int) else None, "center")

        # Col 2: N° de Parte / code
        setc(2, item.get("code",""), cfont(size=10), None, "center")

        # Col 3: Producto / desc (bold)
        setc(3, item.get("desc",""), cfont(size=10, bold=True), None, "left")

        # Col 4: Detalle (italic, gray)
        setc(4, item.get("detalle",""),
             cfont(size=9, italic=True, color="FF666666"), None, "left")

        # Col 5: Estado (colored background)
        estado_val = item.get("estado","")
        estado_key = estado_val.lower()
        e_fill = ESTADO_FILL.get(estado_key)
        e_text = ESTADO_TEXT.get(estado_key, "000000")
        e_bold = ESTADO_BOLD.get(estado_key, False)
        ec = ws.cell(prn, 5)
        ec.value = estado_val
        ec.fill  = sf(e_fill) if e_fill else sf(bg)
        ec.font  = cfont(size=9, bold=e_bold, color="FF"+e_text)
        ec.alignment = Alignment(horizontal="center", vertical="center")
        ec.border = BORDER

        # Col 6: Links (hyperlink if URL)
        ficha = item.get("ficha","")
        lc = ws.cell(prn, 6)
        lc.fill  = sf(bg)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        lc.border = BORDER
        if ficha and (ficha.startswith("http") or ficha.startswith("www")):
            lc.value = "Links"
            lc.font  = cfont(size=9, color="FF0563C1")
            # openpyxl hyperlink
            lc.hyperlink = ficha
        elif ficha:
            lc.value = ficha
            lc.font  = cfont(size=9, color="FF"+GREEN_T)
        else:
            lc.value = ""
            lc.font  = cfont(size=9)

        setc(7,  item.get("qtyCtn",""),    cfont(size=10), None, "center")
        setc(8,  item.get("qtyPlt",""),    cfont(size=10), None, "center")
        setc(9,  item.get("miamiStock",""),cfont(size=10), None, "center")
        setc(10, item.get("miamiEta",""),  cfont(size=10), None, "left")
        setc(11, item.get("bondedStock",""),cfont(size=10), None, "center")
        setc(12, item.get("bondedEta",""), cfont(size=10), None, "left")

        # Col 13: Precio (bold, dark green, $#,##0 sin decimales)
        price = item.get("price")
        if price is not None and price != "":
            price_num = safe_float(price)
            setc(13, price_num,
                 cfont(size=10, bold=True, color="FF"+GREEN_T),
                 "$#,##0", "center")
        else:
            setc(13, "", cfont(size=10), None, "center")

    # ════════════════════════════════════════════════════════
    # Sheet 2 — Cambios
    # ════════════════════════════════════════════════════════
    if changes:
        ws2 = wb.create_sheet("Cambios")
        ws2.merge_cells("A1:E1")
        ws2["A1"] = f"Cambios detectados — {date_str}"
        ws2["A1"].font = cfont(size=13, bold=True, color="FF"+WHITE)
        ws2["A1"].fill = sf("FF6600")
        ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[1].height = 24

        h2 = ["Tipo","N° de Parte","Producto","Valor anterior","Valor nuevo"]
        for col, h in enumerate(h2, 1):
            c = ws2.cell(2, col, value=h)
            c.fill = sf(NAVY); c.font = cfont(10, True, color="FF"+WHITE)
            c.alignment = Alignment(horizontal="center")

        TYPE_CLR = {"new":GREEN_T,"removed":"6B7280","price":"0057B8","stock":"333333"}
        TYPE_LBL = {"new":"Nuevo","removed":"Retirado","price":"Precio","stock":"Stock"}
        for i, ch in enumerate(changes):
            row = i + 3
            bg2 = ALT2 if i % 2 == 0 else ALT1
            t = TYPE_LBL.get(ch.get("type",""), ch.get("type",""))
            old = ch.get("oldPrice", ch.get("oldStock","—"))
            new = ch.get("newPrice", ch.get("newStock","—"))
            if isinstance(old, (int,float)): old = f"${old:,.0f}"
            if isinstance(new, (int,float)): new = f"${new:,.0f}"
            clr = TYPE_CLR.get(ch.get("type",""),"333333")
            for col, val in enumerate([t, ch.get("partNo",""), ch.get("name",""), old, new], 1):
                c = ws2.cell(row, col, value=val)
                c.fill = sf(bg2)
                c.border = BORDER
                c.font = cfont(10, bold=(col==1), color="FF"+clr if col==1 else "FF000000")

        for col, w in enumerate([12,20,46,16,16], 1):
            ws2.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


# ── Build HTML email body ─────────────────────────────────────────────────────
def build_html_email(summary_text, changes, date_str, product_count):
    n_new      = sum(1 for c in changes if c["type"]=="new")
    n_removed  = sum(1 for c in changes if c["type"]=="removed")
    n_price_up = sum(1 for c in changes if c["type"]=="price" and c.get("newPrice",0) > c.get("oldPrice",0))
    n_price_dn = sum(1 for c in changes if c["type"]=="price" and c.get("newPrice",0) < c.get("oldPrice",0))
    n_stock    = sum(1 for c in changes if c["type"]=="stock")
    n_total    = len(changes)

    def pill(label, color, bg):
        return f'<span style="display:inline-block;background:{bg};color:{color};border-radius:4px;padding:2px 8px;font-size:11px;font-weight:700;margin-right:4px">{label}</span>'

    summary_pills = ""
    if n_price_dn: summary_pills += pill(f"▼ {n_price_dn} bajas","#fff","#166534")
    if n_price_up: summary_pills += pill(f"▲ {n_price_up} subas","#fff","#7f1d1d")
    if n_new:      summary_pills += pill(f"+ {n_new} nuevos","#fff","#1e3a5f")
    if n_removed:  summary_pills += pill(f"✕ {n_removed} retirados","#fff","#374151")
    if n_stock:    summary_pills += pill(f"~ {n_stock} stock","#fff","#44337a")

    # Changes table rows
    rows_html = ""
    TYPE_META = {
        "new":     ("#fff","#166534","NUEVO"),
        "removed": ("#fff","#4b5563","RETIRADO"),
        "price":   ("#fff","#1e3a5f","PRECIO"),
        "stock":   ("#fff","#44337a","STOCK"),
    }
    # Show price & new changes first, limit to 50
    priority = sorted(changes, key=lambda c: (0 if c["type"] in ("new","price") else 1, c.get("category","")))
    for i, c in enumerate(priority[:50]):
        meta = TYPE_META.get(c["type"], ("#fff","#555","?"))
        badge = f'<span style="background:{meta[1]};color:{meta[0]};border-radius:3px;padding:1px 6px;font-size:10px;font-weight:700">{meta[2]}</span>'
        bg = "#f9fafb" if i % 2 == 0 else "#ffffff"

        if c["type"] == "price":
            old_v = f'<span style="color:#9ca3af;text-decoration:line-through">${c.get("oldPrice",0):,.2f}</span>'
            new_v = f'<span style="color:#{"166534" if c.get("newPrice",0)<c.get("oldPrice",0) else "7f1d1d"};font-weight:700">${c.get("newPrice",0):,.2f}</span>'
        elif c["type"] == "new":
            old_v = "—"
            new_v = f'<span style="color:#166534;font-weight:700">${c.get("newPrice",0):,.2f}</span>'
        elif c["type"] == "removed":
            old_v = f'${c.get("oldPrice",0):,.2f}'
            new_v = '<span style="color:#9ca3af">—</span>'
        else:  # stock
            old_v = str(c.get("oldStock","—"))
            new_v = str(c.get("newStock","—"))

        rows_html += f"""
        <tr style="background:{bg}">
          <td style="padding:6px 10px;font-size:11px;color:#6b7280">{c.get("partNo","")}</td>
          <td style="padding:6px 10px;font-size:12px;color:#111">{c.get("name","")}</td>
          <td style="padding:6px 10px;font-size:11px;color:#6b7280;text-align:center">{c.get("category","")}</td>
          <td style="padding:6px 10px;font-size:12px;text-align:right">{old_v}</td>
          <td style="padding:6px 10px;font-size:12px;text-align:right">{new_v}</td>
          <td style="padding:6px 10px;text-align:center">{badge}</td>
        </tr>"""

    more = f'<p style="text-align:center;color:#6b7280;font-size:11px">… y {n_total-50} cambios más en el Excel adjunto</p>' if n_total > 50 else ""

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#f3f4f6;font-family:-apple-system,Helvetica,Arial,sans-serif">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#f3f4f6;padding:24px 0">
<tr><td align="center">
<table width="640" cellpadding="0" cellspacing="0" style="max-width:640px;width:100%">

  <!-- Header -->
  <tr><td style="background:#1F2A44;border-radius:10px 10px 0 0;padding:20px 28px">
    <table width="100%"><tr>
      <td><div style="color:#FF6600;font-size:18px;font-weight:700;letter-spacing:-.3px">Alerta de Cambios de Precios</div>
          <div style="color:#8b9ab5;font-size:12px;margin-top:2px">MSI LATAM · SalesOps CRM</div></td>
      <td align="right" style="color:#8b9ab5;font-size:12px;white-space:nowrap">{date_str}</td>
    </tr></table>
  </td></tr>

  <!-- Summary bar -->
  <tr><td style="background:#253047;padding:12px 28px">
    <span style="color:#e2e8f0;font-size:13px">Se detectaron <strong style="color:#fff">{n_total} cambios</strong> en la lista de precios LATAM.
    &nbsp;·&nbsp; {summary_pills}</span>
  </td></tr>

  <!-- Table -->
  <tr><td style="background:#ffffff;padding:0">
    <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse">
      <tr style="background:#1F2A44">
        <th style="padding:8px 10px;font-size:10px;color:#8b9ab5;font-weight:600;text-align:left">CÓDIGO</th>
        <th style="padding:8px 10px;font-size:10px;color:#8b9ab5;font-weight:600;text-align:left">PRODUCTO</th>
        <th style="padding:8px 10px;font-size:10px;color:#8b9ab5;font-weight:600;text-align:center">CAT.</th>
        <th style="padding:8px 10px;font-size:10px;color:#8b9ab5;font-weight:600;text-align:right">PRECIO ANT.</th>
        <th style="padding:8px 10px;font-size:10px;color:#8b9ab5;font-weight:600;text-align:right">PRECIO NUEVO</th>
        <th style="padding:8px 10px;font-size:10px;color:#8b9ab5;font-weight:600;text-align:center">VAR.</th>
      </tr>
      {rows_html}
    </table>
    {more}
  </td></tr>

  <!-- Attachment note -->
  <tr><td style="background:#f8fafc;border:1px solid #e2e8f0;border-top:none;padding:14px 28px">
    <table><tr>
      <td style="font-size:24px;padding-right:12px">📊</td>
      <td>
        <div style="font-weight:600;font-size:13px;color:#1e293b">MSI LATAM - Lista de Precios - {date_str.replace("/",".")}.xlsx</div>
        <div style="font-size:11px;color:#64748b">Lista completa con {product_count} productos · Hojas: Lista de Precios + Cambios</div>
      </td>
    </tr></table>
  </td></tr>

  <!-- Footer -->
  <tr><td style="background:#1F2A44;border-radius:0 0 10px 10px;padding:14px 28px;text-align:center">
    <span style="color:#8b9ab5;font-size:11px">Enviado automáticamente por MSI SalesOps CRM ·
    <a href="https://msicrm.com" style="color:#FF6600;text-decoration:none">msicrm.com</a></span>
  </td></tr>

</table>
</td></tr></table>
</body></html>"""
    return html


# ── Send email ────────────────────────────────────────────────────────────────
def send_email(summary_text, changes, excel_bytes, date_str, product_count):
    if not RECIPIENTS:
        print("[SKIP] No recipients configured"); return
    if not SMTP_PASS:
        print("[SKIP] SMTP_PASSWORD not set"); return

    msg = MIMEMultipart("alternative")
    msg["From"]         = f"MSI Argentina <{SMTP_USER}>"
    msg["To"]           = SMTP_USER
    msg["Bcc"]          = ", ".join(RECIPIENTS)
    msg["Subject"]      = f"[MSI CRM] Cambios en lista de precios LATAM — {date_str}"
    msg["Reply-To"]     = SMTP_USER
    msg["X-Mailer"]     = "MSI SalesOps CRM"
    msg["Precedence"]   = "bulk"
    msg["Auto-Submitted"] = "auto-generated"

    # Plain text fallback
    plain = f"MSI LATAM — Cambios en lista de precios\n{date_str}\n\n{summary_text}\n\nVer Excel adjunto.\n\nsales@msicrm.com"
    msg.attach(MIMEText(plain, "plain", "utf-8"))

    # HTML body
    html_body = build_html_email(summary_text, changes, date_str, product_count)
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    # Excel attachment — needs a new MIMEMultipart("mixed") wrapper
    outer = MIMEMultipart("mixed")
    outer["From"]         = msg["From"]
    outer["To"]           = msg["To"]
    outer["Bcc"]          = msg["Bcc"]
    outer["Subject"]      = msg["Subject"]
    outer["Reply-To"]     = msg["Reply-To"]
    outer["X-Mailer"]     = msg["X-Mailer"]
    outer["Precedence"]   = msg["Precedence"]
    outer["Auto-Submitted"] = msg["Auto-Submitted"]
    outer.attach(msg)

    if excel_bytes:
        filename = f"MSI LATAM - Lista de Precios - {date_str.replace('/','.') }.xlsx"
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(excel_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        outer.attach(part)

    try:
        with smtplib.SMTP_SSL(SMTP_HOST, SMTP_PORT) as server:
            server.login(SMTP_USER, SMTP_PASS)
            server.sendmail(SMTP_USER, RECIPIENTS, outer.as_bytes())
        print(f"[OK]   Email enviado a {len(RECIPIENTS)} destinatario(s)")
    except Exception as e:
        print(f"[ERR]  Email failed: {e}")


# ── Main ──────────────────────────────────────────────────────────────────────
print("=== MSI Stock & Price Check ===")
now_iso  = datetime.now(timezone.utc).isoformat()
date_str = datetime.now().strftime("%d/%m/%Y")

# 1. Load pricelist
print("Loading pricelist from Firebase...")
pricelist_raw = fb_get("salesops_pricelist")
if not pricelist_raw:
    local_path = os.path.join(os.path.dirname(__file__), '..', 'pricelist_snapshot.json')
    if os.path.exists(local_path):
        with open(local_path) as f: pricelist_raw = json.load(f)
        print("Fallback: local pricelist_snapshot.json")
    else:
        print("[ERROR] No pricelist available. Exiting."); sys.exit(1)

if isinstance(pricelist_raw, str): pricelist_raw = json.loads(pricelist_raw)
raw_products = pricelist_raw.get("products", {})

# Firebase may return products as list — normalize to dict
if isinstance(raw_products, list):
    raw_products = {
        (p.get("code") or p.get("partNo") or p.get("sku") or str(i)): p
        for i, p in enumerate(raw_products) if isinstance(p, dict)
    }
print(f"Pricelist: {len(raw_products)} products")

# 2. Build snapshot (lean, for diffing)
# Field names match the CRM: code, desc, price (not name/partNo)
new_snapshot = {}
items_for_snap = raw_products if isinstance(raw_products, dict) else {
    (p.get("code") or str(i)): p for i, p in enumerate(raw_products) if isinstance(p, dict)
}
for key, p in items_for_snap.items():
    if p.get("isSubSection"): continue   # skip section markers
    code = p.get("code") or key
    new_snapshot[code] = {
        "name":        p.get("desc", p.get("name","")),
        "category":    p.get("category",""),
        "miamiStock":  safe_int(p.get("miamiStock",0)),
        "miamiPrice":  safe_float(p.get("price", p.get("priceR", 0))),
        "miamiEta":    p.get("miamiEta",""),
        "bondedStock": safe_int(p.get("bondedStock",0)),
    }

# 3. Load old snapshot
old_raw = fb_get("salesops_stock_snapshot")
if old_raw and isinstance(old_raw, str): old_raw = json.loads(old_raw)
old_snapshot = old_raw if isinstance(old_raw, dict) else {}
print(f"Previous snapshot: {len(old_snapshot)} products")

# 4. Compute diff
changes = []
for partNo in set(new_snapshot) | set(old_snapshot):
    n = new_snapshot.get(partNo)
    o = old_snapshot.get(partNo)
    if n and not o:
        changes.append({"type":"new","partNo":partNo,"name":n["name"],"category":n["category"],"newPrice":n["miamiPrice"],"newStock":n["miamiStock"]})
    elif o and not n:
        changes.append({"type":"removed","partNo":partNo,"name":o["name"],"category":o["category"],"oldPrice":o["miamiPrice"]})
    elif n and o:
        if abs(n["miamiPrice"] - o["miamiPrice"]) >= 0.5:
            changes.append({"type":"price","partNo":partNo,"name":n["name"],"category":n["category"],"oldPrice":o["miamiPrice"],"newPrice":n["miamiPrice"]})
        if abs(n["miamiStock"] - o["miamiStock"]) >= 1:
            changes.append({"type":"stock","partNo":partNo,"name":n["name"],"category":n["category"],"oldStock":o["miamiStock"],"newStock":n["miamiStock"]})

has_changes = len(changes) > 0
summary = f"{len(changes)} cambio(s) detectado(s)" if has_changes else "Sin cambios"
print(f"Changes: {summary}")

# 5. Write to Firebase
diff_payload = {
    "checkedAt": now_iso, "productCount": len(new_snapshot),
    "hasChanges": has_changes, "changeCount": len(changes),
    "changes": changes[:100], "summary": summary,
    "triggeredBy": "github-actions", "triggerMode": "auto",
}
changelog_entry = {
    "checkedAt": now_iso, "productCount": len(new_snapshot),
    "hasChanges": has_changes, "changeCount": len(changes),
    "summary": summary, "triggerMode": "auto", "triggeredBy": "github-actions",
}

print("Writing to Firebase...")
ok1 = fb_put("salesops_stock_diff",       json.dumps(diff_payload))
ok2 = fb_put("salesops_stock_snapshot",   json.dumps(new_snapshot))
ok3 = fb_post("salesops_stock_changelog", changelog_entry)

# 6. Update local snapshot files
repo_root = os.path.join(os.path.dirname(__file__), '..')
with open(os.path.join(repo_root, 'stock_diff_latest.json'), 'w') as f:
    json.dump(diff_payload, f, ensure_ascii=False, indent=2)
with open(os.path.join(repo_root, 'stock_snapshot_latest.json'), 'w') as f:
    json.dump(new_snapshot, f, ensure_ascii=False, indent=2)

# 7. Send email
if has_changes or FORCE_EMAIL:
    if FORCE_EMAIL and not has_changes:
        summary = f"Envío de prueba — {len(new_snapshot)} productos en lista"
        print("FORCE_EMAIL=true — sending regardless of changes")
    else:
        print("Generating Excel and sending email...")
    excel_bytes = generate_excel(raw_products, changes, date_str)
    send_email(summary, changes, excel_bytes, date_str, len(new_snapshot))
else:
    print("No changes — skipping email")

print(f"\n{'='*40}")
print(f"RESULT: {summary}")
print(f"Firebase: diff={'OK' if ok1 else 'ERR'} snapshot={'OK' if ok2 else 'ERR'} changelog={'OK' if ok3 else 'ERR'}")
if not (ok1 and ok2 and ok3): sys.exit(1)
